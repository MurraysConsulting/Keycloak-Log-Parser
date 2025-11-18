# Import support
import pandas as pd
import os
import glob
from datetime import datetime as dt
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import string
from collections import defaultdict


# Main function to orchestrate all the functions below in a managed way.
def main() -> None:
    # Set data type loads for inputting data into parser. This could be adapted in the future to be dynamic,
    # and ideally loaded to a DB to reduce having to load old data every run.
    data_folder = '\\Data\\'  # This is where all the data is located for storing the raw logs and user data on the system.
    new_log_load = '*block*.json'  # This is to find the files for any new logs sent from RK to be parsed.
    old_log_load = '*keycloak*.json'  # This is to find the files for older logs prior to the update to keycloak
    sys_user_load = 'response*.json'  # This is to find the file for loading system user information from RK Keycloak
    request_user_load = '*.csv'  # This is to find the latest user account request information that has been processed by the help desk for RK accounts.
    wbname="Main_User_Info.xlsx"
    wbexport="test2.xlsx"
    

    raw_log_data = []
    log_data = []  # This is where we are going to load the data from the external logs and where it will be parsed.
    
    userlist_data = {'sys': [], 'req': [], 'log': []}  # Creates a dictionary with two lists one for system users, and one for requested user accounts list
    # The log entry on userlist_data is to show all the users found in the log data info.

    missing_user_data = {'sys': [], 'req': []}  # Creates a dictionary with two lists of missing users when comparing the list from the system and requested accounts.

    # Let's load from new data sources, then we can add from the historical data sources
    raw_log_data = import_json_files(data_folder, new_log_load)  # New data log format loads
    print('loading done')
    raw_log_data = new_json_parse(raw_log_data)
    print('load')
    log_data.extend(log_data_parse(raw_log_data))
    print('end of first call')
    
    raw_log_data = import_json_files(data_folder, old_log_load)  # Old data log format loads
    print('loading done')
    log_data.extend(log_data_parse(raw_log_data))
    print('end of second call')

    # Load the info from the supplied files for users as recorded in the system, and as recorded in Jira for requested accounts.
    userlist_data['sys'] = userList_parse(import_json_files(data_folder, sys_user_load))
    userlist_data['req'] = import_csv_files(data_folder, request_user_load)

    # Find users that are missing from the different user lists.
    missing_user_data['sys'], missing_user_data['req'], userlist_data['sys'] = compare_sys_req(userlist_data['sys'], userlist_data['req'])
    
    # Add user orgs to the log data for better parsing and association of data.
    log_data = add_org(log_data, userlist_data['sys'], missing_user_data['sys'])

    # Debug output
    for data in log_data[:5]:  # Print first 5 entries
        print(data)

    print("\nUser List Data:")
    print(f"System users: {len(userlist_data['sys'])}")
    print(f"Requested users: {len(userlist_data['req'])}")

    print("\nMissing User Data:")
    print(f"System users missing: {len(missing_user_data['sys'])}")
    print(f"Requested users missing: {len(missing_user_data['req'])}")

    wb, logsheet = loadexcel(wbname)

    logsheet = load_log_data(log_data, wb, logsheet)

    # Generate analytics sheets
    print("\nGenerating analytics...")
    generate_user_events_by_month(log_data, wb)
    generate_users_per_org_by_month(log_data, wb)
    generate_users_per_service_by_month(log_data, wb)

    export_excel(wb, wbexport)
    print(f"\nExport complete: {wbexport}")


# NEW ANALYTICS FUNCTIONS

def generate_user_events_by_month(log_data, workbook):
    """Generate sheet showing unique users with number of events per month"""
    print("Generating user events by month...")
    
    # Structure: {username: {(year, month): event_count}}
    user_events = defaultdict(lambda: defaultdict(int))
    
    for entry in log_data:
        if 'username' in entry and 'year' in entry and 'month' in entry:
            username = entry['username']
            year_month = (entry['year'], entry['month'])
            user_events[username][year_month] += 1
    
    # Convert to list of dicts for DataFrame
    data_rows = []
    for username, months in user_events.items():
        for (year, month), count in sorted(months.items()):
            data_rows.append({
                'Username': username,
                'Year': year,
                'Month': month,
                'Event_Count': count
            })
    
    if data_rows:
        df = pd.DataFrame(data_rows)
        create_excel_table(workbook, df, 'User_Events_By_Month', 'UserEventsTable')
    else:
        print("No user event data to generate")


def generate_users_per_org_by_month(log_data, workbook):
    """Generate sheet showing number of users per organization per month"""
    print("Generating users per organization by month...")
    
    # Structure: {(org, year, month): set(usernames)}
    org_users = defaultdict(set)
    
    for entry in log_data:
        if 'org' in entry and 'username' in entry and 'year' in entry and 'month' in entry:
            org = entry['org']
            username = entry['username']
            year_month = (org, entry['year'], entry['month'])
            org_users[year_month].add(username)
    
    # Convert to list of dicts for DataFrame
    data_rows = []
    for (org, year, month), users in sorted(org_users.items()):
        data_rows.append({
            'Organization': org,
            'Year': year,
            'Month': month,
            'Unique_User_Count': len(users)
        })
    
    if data_rows:
        df = pd.DataFrame(data_rows)
        create_excel_table(workbook, df, 'Users_Per_Org_By_Month', 'OrgUsersTable')
    else:
        print("No organization data to generate")


def generate_users_per_service_by_month(log_data, workbook):
    """Generate sheet showing number of unique users per system service per month"""
    print("Generating users per service by month...")
    
    # Structure: {(clientId, year, month): set(usernames)}
    service_users = defaultdict(set)
    
    for entry in log_data:
        if 'clientId' in entry and 'username' in entry and 'year' in entry and 'month' in entry:
            service = entry['clientId']
            username = entry['username']
            year_month = (service, entry['year'], entry['month'])
            service_users[year_month].add(username)
    
    # Convert to list of dicts for DataFrame
    data_rows = []
    for (service, year, month), users in sorted(service_users.items()):
        data_rows.append({
            'Service': service,
            'Year': year,
            'Month': month,
            'Unique_User_Count': len(users)
        })
    
    if data_rows:
        df = pd.DataFrame(data_rows)
        create_excel_table(workbook, df, 'Users_Per_Service_By_Month', 'ServiceUsersTable')
    else:
        print("No service data to generate")


def create_excel_table(workbook, dataframe, sheet_name, table_name):
    """Helper function to create a formatted Excel table from a DataFrame"""
    # Remove sheet if it exists
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    
    sheet = workbook.create_sheet(sheet_name)
    
    # Load the dataframe data into the sheet
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        sheet.append(row)
    
    # Get sheet dimensions
    max_column = sheet.max_column
    max_row = sheet.max_row
    
    # Calculate column letter for table
    abc = list(string.ascii_lowercase)
    tablecolumn = max_column
    extra_letter = ''
    
    if tablecolumn > 26:
        extra_letter = abc[(tablecolumn // 26) - 1]
        tablecolumn = tablecolumn % 26
        if tablecolumn == 0:
            tablecolumn = 26
    
    column_letter = extra_letter + abc[tablecolumn - 1]
    tablesize = f'A1:{column_letter}{max_row}'
    
    # Create and style the table
    tab = Table(displayName=table_name, ref=tablesize)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    
    return sheet


# EXISTING FUNCTIONS (unchanged)

def import_json_files(folder, tgt):
    destination = os.path.join(os.getcwd(), folder.strip('\\'))
    
    file_list = glob.glob(os.path.join(destination, tgt))  # Will collect all files in the destination with the tgt's name or group
    
    files_content = []
    
    for file in file_list:
        try:
            with open(file, 'r') as f:
                content = json.load(f)
                if isinstance(content, list):
                    files_content.extend(content)
                else:
                    files_content.append(content)
        except json.JSONDecodeError as e:
            print(f"Failed: {file}")
            print(f"JSON Error: {e}")
            print("Looks like it is not a valid json format. Go to https://jsonlint.com/ to understand how to fix the error")
        except Exception as e:
            print(f"Failed: {file}")
            print(f"Error: {e}")
    
    return files_content


def import_csv_files(folder, tgt):
    destination = os.path.join(os.getcwd(), folder.strip('\\'))
    
    # Get the first CSV file that matches the pattern
    file_list = glob.glob(os.path.join(destination, tgt))
    
    if not file_list:
        print(f"No CSV files found matching pattern: {tgt}")
        return []
    
    csv_file = file_list[0]  # Use the first matching file
    
    csvKey = []
    outputList = []
    
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            for count, line in enumerate(f):
                line = line.strip('\n').strip('"')
                
                if count == 0:
                    # Header row
                    csvKey = [col.strip('"') for col in line.split('","')]
                else:
                    # Data rows
                    values = [val.strip('"') for val in line.split('","')]
                    if len(values) == len(csvKey):
                        outputList.append(dict(zip(csvKey, values)))
    except Exception as e:
        print(f"Error reading CSV file {csv_file}: {e}")
    
    return outputList


def new_json_parse(data):
    parsedData = []
    
    for item in data:
        try:
            if "records" in item:
                parsedData += item["records"]

        except Exception as e:
            print(f'Error parsing item: {e}')
    
    print(f"Parsed {len(parsedData)} records from new format")
    return parsedData


def userList_parse(userlist):
    if not userlist:
        return []
    
    # Handle if userlist is a list with a single dict containing 'data'
    if isinstance(userlist, list) and len(userlist) > 0:
        if isinstance(userlist[0], dict) and 'data' in userlist[0]:
            userlist = userlist[0]
    
    updatedUserList = []
    
    try:
        if isinstance(userlist, dict) and 'data' in userlist and 'Users' in userlist['data']:
            for user in userlist['data']['Users']:
                updatedUserList.append(user)
    except Exception as e:
        print(f"Error parsing user list: {e}")
    
    return updatedUserList


def log_data_parse(data):
    # Function variables for parsing
    data_output = []  # Output of the cleaned log data
    dataKeys = ['type', 'clientId', 'userId', 'username', 'identity_provider_identity']
    idstore = set()  # Use set for faster lookups
    
    print('working 1')
    count = 0

    for entry in data:
        tempLog = {}

        try:
            if '@message' in entry.keys():
                
                # Creates a unique key for the data entry
                tempLog['id'] = entry['@message']['json']['time']
                
                # Pulls out the log entry that we are looking for to parse
                logitem = entry['@message']['json']['log'].split(',')
                
                tempDate, tempTime = entry['@message']['json']['time'].split('T')
                
                # Breaks down the entry info into smaller pieces for ease of filtering by individual fields
                tempLog['year'] = int(tempDate.split('-')[0])
                tempLog['month'] = int(tempDate.split('-')[1])
                tempLog['day'] = int(tempDate.split('-')[2])
                tempLog['time'] = tempTime.split('Z')[0]

                for item in logitem:
                    if '=' in item:
                        item_parts = item.split('=')
                        key = item_parts[0].split(' ')[-1]
                        value = item_parts[1] if len(item_parts) > 1 else ''

                        if key in dataKeys:
                            tempLog[key] = value
                
                if 'type' in tempLog.keys():
                    if tempLog['id'] not in idstore:  # Removes duplicate entries
                        idstore.add(tempLog['id'])
                        data_output.append(tempLog)
                count += 1
        except Exception as e:
            print(f"Error parsing user list: {e}")

    # Clean the keys of any extra characters and clean up the name of the service
    userInfo = []
    print('working')

    for entry in data_output:
        for key in entry:
            if isinstance(entry[key], str):
                if '"' in entry[key]:
                    entry[key] = entry[key].strip('"')
                    
                if key == 'clientId':
                    if 'https' in entry[key]:
                        entry[key] = 'metadata'
        
        if 'userId' and 'username' in entry:
            tempUser = {'userId': entry['userId'], 'username': entry['username']}
            if tempUser not in userInfo:
                userInfo.append(tempUser)
    
    # Create a lookup dictionary for faster username retrieval
    user_lookup = {user['userId']: user['username'] for user in userInfo}
    
    # Adding username for entries that only have userId
    for entry in data_output:
        if 'username' not in entry and 'userId' in entry:
            if entry['userId'] in user_lookup:
                entry['username'] = user_lookup[entry['userId']]
    
    return data_output


def compare_sys_req(list1, list2):
    syslist = list1
    reqlist = list2
    sysusermissing = []
    requsermissing = []
    
    # Create a set of request emails for faster lookup
    req_emails = {req['Email'] for req in reqlist if 'Email' in req}
    
    # Check system users against request list
    for sysuser in syslist:
        if 'email' not in sysuser:
            continue
            
        if sysuser['email'] in req_emails:
            # Find the matching request and add org
            for requser in reqlist:
                if 'Email' in requser and requser['Email'] == sysuser['email']:
                    if 'RK-Organization' in requser:
                        sysuser['org'] = requser['RK-Organization']
                    break
        else:
            if 'username' in sysuser:
                sysusermissing.append(sysuser['username'])
    
    # Create a set of system emails for faster lookup
    sys_emails = {sys['email'] for sys in syslist if 'email' in sys}
    
    # Check request users against system list
    for requser in reqlist:
        if 'Email' in requser:
            if requser['Email'] not in sys_emails:
                requsermissing.append(requser['Email'])
    
    return sysusermissing, requsermissing, syslist


def add_org(data, ulist, sumissing):
    log_data = data
    userlist = ulist
    sysusermissing = set(sumissing)  # Convert to set for faster lookup
    
    # Create a lookup dictionary for faster user retrieval
    user_lookup = {user['id']: user for user in userlist if 'id' in user}
    
    for entry in log_data:
        if 'userId' not in entry:
            continue
            
        user_id = entry['userId']
        
        if user_id in user_lookup:
            user = user_lookup[user_id]
            
            # Set RK-Org if available
            if 'attributes' in user and 'gov-organization' in user['attributes']:
                entry['RK-Org'] = user['attributes']['gov-organization']
            
            # Set org based on whether user is in missing list
            if 'username' in user:
                if user['username'] not in sysusermissing:
                    # Sets the org to what has been defined as the users owning org per the request for account ticket
                    if 'org' in user:
                        entry['org'] = user['org']
                else:
                    # Sets org to vendor that is developing/maintaining the sys. Backend User Accounts.
                    entry['org'] = 'BAH-Developer'
    
    return log_data


def loadexcel(fname):
    filename = fname
    
    # Load file
    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()
        # Remove default sheet if creating new workbook
        if 'Sheet' in workbook.sheetnames:
            workbook.create_sheet('logs')
            workbook.remove(workbook['Sheet'])
    
    # Check if 'logs' sheet exists and remove it
    if 'logs' in workbook.sheetnames:
        logsheet = workbook['logs']
        workbook.remove(logsheet)
    
    logsheet = workbook.create_sheet('logs')
    
    return workbook, logsheet


def load_log_data(data, workbook, sheet):
    logsheet = sheet
    logworkbook = workbook
    df_loginfo = pd.DataFrame(data)  # Takes input data and moves it into a data frame for easy excel sheet builds.
    
    # Load the dataframe data into the blank sheet.
    for row in dataframe_to_rows(df_loginfo, index=False, header=True):
        logsheet.append(row)
    
    # Understanding the sheet to be able to build a table for the data just loaded
    maxColumn = logsheet.max_column
    maxRow = logsheet.max_row
    
    # List of letters 0-25 for the alphabet for understanding excel columns
    abc = list(string.ascii_lowercase)
    
    # Calculate column letter for table
    tablecolumn = maxColumn
    extra_letter = ''
    
    if tablecolumn > 26:
        extra_letter = abc[(tablecolumn // 26) - 1]
        tablecolumn = tablecolumn % 26
        if tablecolumn == 0:
            tablecolumn = 26
    
    column_letter = extra_letter + abc[tablecolumn - 1]
    tablesize = f'A1:{column_letter}{maxRow}'
    
    # Define the name of the table for use in the logsheet to encompass the data that was just loaded above
    tab = Table(displayName="Table1", ref=tablesize)
    
    # Sets the style to be applied to the table when it is created
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    
    # This will take all the work above and store it for use on the logsheet
    tab.tableStyleInfo = style
    
    # Takes all the prep work and creates the table around the data just loaded into the logsheet.
    logsheet.add_table(tab)
    
    return logworkbook


def export_excel(workbook, fname):
    workbook.save(fname)


if __name__ == "__main__":
    main()
